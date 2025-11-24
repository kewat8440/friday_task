import configparser
import psycopg2
import openpyxl
import os

# ---------------- DB CONFIG ---------------- #

config = configparser.ConfigParser()
config.read('config.ini')

db_username = config['DB_Credentials']['username']
db_password = config['DB_Credentials']['password']
db_host     = config['DB_Credentials']['host']
db_port     = config['DB_Credentials']['port']
db_name     = config['DB_Credentials']['db_name']


def get_connection():
    """Create and return a PostgreSQL connection using config.ini."""
    return psycopg2.connect(
        host=db_host,
        port=db_port,
        user=db_username,
        password=db_password,
        dbname=db_name
    )


# --------------- CORE FUNCTIONS ---------------- #

def header(input_sheet):
    """
    Read header from row 5 (A5:H5), clean it, and build SQL column definition.
    Adjust A5:H5 if you want more/less columns.
    """
    head = [cell.value for cell in input_sheet['A5':'H5'][0]]
    print("Raw header row:", head)

    # Clean: convert to string, strip spaces, replace spaces with underscore
    headers = [str(h).strip().replace(" ", "_") for h in head]
    columns_sql = ", ".join([f'"{c}" TEXT' for c in headers])

    print("Final headers for SQL:", columns_sql)
    return headers, columns_sql


def create_table(cols):
    """
    Create dam_market_data table if it doesn't exist.
    All Excel columns stored as TEXT.
    """
    create_sql = f"""
    CREATE TABLE IF NOT EXISTS dam_market_data (
        id SERIAL PRIMARY KEY,
        {cols}
    );
    """

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute(create_sql)
        conn.commit()
        cursor.close()
        conn.close()
        print("Table ensured: dam_market_data")
    except Exception as e:
        print("Error creating table:", e)


def all_row_insert_into_database(input_sheet, header_list):
    """
    Insert all data rows starting from row 6 until the last row.
    Skips fully empty rows.
    """
    try:
        conn = get_connection()
        cursor = conn.cursor()

        col_names_str = ", ".join([f'"{c}"' for c in header_list])

        start_row = 6
        end_row = input_sheet.max_row
        max_col = len(header_list)

        insert_sql = f"""
        INSERT INTO dam_market_data ({col_names_str})
        VALUES ({", ".join(["%s"] * max_col)})
        """

        for row in input_sheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=1,
            max_col=max_col,
            values_only=True
        ):
            # Skip completely empty rows
            if row is None or all(cell is None for cell in row):
                continue

            cursor.execute(insert_sql, row)

        conn.commit()
        cursor.close()
        conn.close()
        print("All data inserted")
    except Exception as e:
        print("Error inserting data:", e)


def check_excel_in_folder(folder_path):
    """
    Check folder for Excel files and process them.
    Returns True if any Excel file found, else False.
    """
    excel_extensions = ('.xls', '.xlsx', '.xlsm')

    try:
        files = os.listdir(folder_path)
    except FileNotFoundError:
        print(f"Folder does not exist: {folder_path}")
        return False

    excel_files = [
        file for file in files
        if file.lower().endswith(excel_extensions)
    ]

    if not excel_files:
        return False

    print("Excel files found:", excel_files)

    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        print("\nProcessing:", file)

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        header_list, columns_sql = header(sheet)

        create_table(columns_sql)
        all_row_insert_into_database(sheet, header_list)

        wb.close()

    return True


# --------------- BOT WRAPPER ---------------- #

def excel_bot(folder_path: str):
    """
    Bot behavior:
      - If folder has no Excel files -> just exit.
      - If Excel files exist -> perform DB logic.
    """
    if not os.path.exists(folder_path):
        print(f"Folder does not exist: {folder_path}")
        return

    excel_found = check_excel_in_folder(folder_path)

    if not excel_found:
        print("No Excel file found in the folder. Bot exiting.")
        return

    print("Bot work completed successfully.")


# --------------- MAIN ---------------- #

if __name__ == "__main__":
    # You can also move this folder path to config.ini if you want.
    folder_path = r'/Users/santoshkewat/Desktop/fridadytask/friday_task/Excel Automation '
    excel_bot(folder_path)
